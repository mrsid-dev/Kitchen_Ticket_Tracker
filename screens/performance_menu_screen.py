from datetime import timedelta, datetime
from kivy.uix.screenmanager import Screen
from kivy.uix.scrollview import ScrollView
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.popup import Popup
import sqlite3, openpyxl, os, pytz
from datetime import datetime, timedelta
from openpyxl.styles import Font
from plyer import storagepath
from pathlib import Path
from utils.customboxlayouts import RoundedBoxLayout, RoundedButton, ColoredBoxLayout
from db.db_initialization import db_path  # ✅ Import the correct path

def convert_utc_to_local(utc_timestamp):
    """Convert a UTC timestamp to local (EST/EDT) time."""
    eastern = pytz.timezone("US/Eastern")
    utc_dt = datetime.strptime(utc_timestamp, "%Y-%m-%d %H:%M:%S").replace(tzinfo=pytz.utc)
    return utc_dt.astimezone(eastern)  # ✅ Now it returns the converted timestamp


class PerformanceMenuScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.current_view = ''

        # Main Layout
        main_layout = ColoredBoxLayout(orientation="vertical", spacing=10, padding=20, color=(0.118, 0.231, 0.208, 1))
        self.add_widget(main_layout)

        # Header (Title)
        header = BoxLayout(size_hint=(1, None), height=50)
        self.title_label = Label(text="Performance Data", font_size=60, size_hint=(1, None), height=40, bold=True, underline=True)
        header.add_widget(self.title_label)
        main_layout.add_widget(header)

        # Scrollable Middle Content
        scrollable_container = BoxLayout(orientation="vertical", size_hint=(1, 1), padding=10)
        scroll_view = ScrollView(size_hint=(1, 1), do_scroll_x=False, do_scroll_y=True)
        self.data_container = BoxLayout(orientation="vertical", spacing=10, size_hint_y=None)
        self.data_container.bind(minimum_height=self.data_container.setter("height"))
        scroll_view.add_widget(self.data_container)
        scrollable_container.add_widget(scroll_view)
        main_layout.add_widget(scrollable_container)

        # Footer (Buttons)
        footer = BoxLayout(orientation="horizontal", size_hint=(1, None), height=150, spacing=10)
        view_selection_toggle_container = BoxLayout(orientation="horizontal", size_hint=(1, None), height=150, spacing=0)
        view_selection_toggle = RoundedBoxLayout(orientation="horizontal", size_hint=(1, None), height=150, spacing=10, padding=[8, 0, 8, 0], color=(0.204, 0.408, 0.373, 1))

        self.monthly_button = RoundedButton(text="[size=24][b]Month[/b][/size]\n[size=60][font=fonts/MaterialIcons-Regular.ttf]\uefe7[/font][/size]",
                                            size_hint=(0.25, None), height=130,
                                            on_press=self.load_monthly_data,
                                            background_color=(0.204, 0.408, 0.373, 1),
                                            color=(0.714, 0.569, 0.129, 1),
                                            markup=True)
        self.monthly_button.pos_hint = {'center_y': 0.5, 'center_x': 0.5}
        view_selection_toggle.add_widget(self.monthly_button)

        self.weekly_button = RoundedButton(text="[size=24][b]Week[/b][/size]\n[size=62][font=fonts/MaterialIcons-Regular.ttf]\uefe8[/font][/size]",
                                           size_hint=(0.25, None), height=130,
                                           on_press=self.load_weekly_data,
                                           background_color=(0.204, 0.408, 0.373, 1),
                                           color=(0.714, 0.569, 0.129, 1),
                                           markup=True)
        self.weekly_button.pos_hint = {'center_y': 0.5, 'center_x': 0.5}
        view_selection_toggle.add_widget(self.weekly_button)

        self.daily_button = RoundedButton(text="  [size=24][b]Day[/b][/size]\n[size=62][font=fonts/MaterialIcons-Regular.ttf]\ue8ed[/font][/size]",
                                          size_hint=(0.25, None), height=130,
                                          on_press=self.load_daily_data,
                                          background_color=(0.204, 0.408, 0.373, 1),
                                          color=(0.714, 0.569, 0.129, 1),
                                          markup=True)
        self.daily_button.pos_hint = {'center_y': 0.5, 'center_x': 0.5}
        view_selection_toggle.add_widget(self.daily_button)

        self.hourly_button = RoundedButton(
            text="[size=24][b]Hour[/b][/size]\n[size=62][font=fonts/MaterialIcons-Regular.ttf]\ue8e9[/font][/size]",
            size_hint=(0.25, None), height=130,
            on_press=self.load_hourly_data,
            background_color=(0.204, 0.408, 0.373, 1),
            color=(0.714, 0.569, 0.129, 1),
            markup=True)
        self.hourly_button.pos_hint = {'center_y': 0.5, 'center_x': 0.5}
        view_selection_toggle.add_widget(self.hourly_button)

        view_selection_toggle_container.add_widget(view_selection_toggle)
        footer.add_widget(view_selection_toggle_container)

        buttons = BoxLayout(orientation="horizontal", size_hint=(1, None), height=150, spacing=10)

        self.back_button = RoundedButton(
            text="[font=fonts/MaterialIcons-Regular.ttf][size=45]\ue2ea[/size][/font] [size=40][b]Back[/b][/size]",
            size_hint_y=None,
            height=150,
            background_normal="",
            background_color=(0.541, 0.29, 0.29, 1),  # Bright red (#c33f42)
            markup=True,
        )
        self.back_button.bind(on_press=self.go_back)

        # Add Export button
        self.export_button = RoundedButton(
            text="[font=fonts/MaterialIcons-Regular.ttf][size=60]\ue9a3[/size][/font] [size=40][b]Export[/b][/size]",
            size_hint_y=None,
            height=150,
            background_normal="",
            background_color=(0.373, 0.392, 0.408, 1,),
            markup=True
        )
        self.export_button.bind(on_press=self.export_to_excel)

        # Add buttons to the footer
        main_layout.add_widget(footer)
        buttons.add_widget(self.export_button)
        buttons.add_widget(self.back_button)
        main_layout.add_widget(buttons)

    def _update_bg_rect(self, *args):
        """Update background rectangle size and position."""
        self.bg_rect.size = self.size
        self.bg_rect.pos = self.pos

    def load_monthly_data(self, instance):
        self.current_view = "monthly"
        # Set button states correctly
        self.monthly_button.set_active(True)
        self.weekly_button.set_active(False)
        self.daily_button.set_active(False)
        self.hourly_button.set_active(False)
        self.load_performance_data(group_by="month")

    def load_weekly_data(self, instance):
        self.current_view = "weekly"
        # Set button states correctly
        self.monthly_button.set_active(False)
        self.weekly_button.set_active(True)
        self.daily_button.set_active(False)
        self.hourly_button.set_active(False)
        self.load_performance_data(group_by="week")

    def load_daily_data(self, instance):
        self.current_view = "daily"
        # Set button states correctly
        self.monthly_button.set_active(False)
        self.weekly_button.set_active(False)
        self.daily_button.set_active(True)
        self.hourly_button.set_active(False)
        self.load_performance_data(group_by="day")

    def load_hourly_data(self, instance):
        self.current_view = "hourly"
        # Set button states correctly
        self.monthly_button.set_active(False)
        self.weekly_button.set_active(False)
        self.daily_button.set_active(False)
        self.hourly_button.set_active(True)
        self.load_performance_data(group_by="hour")

    def load_performance_data(self, group_by="day"):
        """Fetch and display performance data grouped by the specified period."""
        self.data_container.clear_widgets()

        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        # Get current local time and determine start_date based on selected range
        local_tz = pytz.timezone("US/Eastern")  # Change to your actual timezone
        now = datetime.now(local_tz)

        if group_by == "hour":
            start_date = now - timedelta(days=2)  # Last 2 days of hourly data
            date_format = "%Y-%m-%d %H:00"
        elif group_by == "day":
            start_date = now - timedelta(days=7)  # Last 7 days
            date_format = "%Y-%m-%d"
        elif group_by == "week":
            start_date = now - timedelta(weeks=4)  # Last 4 weeks
            date_format = "%Y-%W"
        elif group_by == "month":
            start_date = now - timedelta(days=90)  # Last 3 months
            date_format = "%Y-%m"
        else:
            return

        # Convert `start_date` to UTC before filtering the database
        start_date_utc = start_date.astimezone(pytz.utc).strftime("%Y-%m-%d %H:%M:%S")

        print(f"Filtering for: {group_by} | Start Date (Local): {start_date} | Start Date (UTC): {start_date_utc}")

        # Query to pull the correct date range (convert UTC timestamps to local)
        cursor.execute(f"""
            SELECT 
                date AS utc_timestamp,
                cooks.name AS cook_name,
                tickets.time_taken
            FROM tickets
            INNER JOIN cooks ON cooks.pin = tickets.cook_pin
            WHERE date >= ?
            ORDER BY utc_timestamp ASC, cook_name ASC
        """, (start_date_utc,))

        results = cursor.fetchall()
        conn.close()

        if not results:
            self.data_container.add_widget(Label(text="No data found!"))  # Show this if no data
            return

        # Process data
        aggregated_data = {}
        for utc_timestamp, cook_name, time_taken in results:
            local_dt = convert_utc_to_local(utc_timestamp)  # Convert from UTC → EST/EDT
            period = local_dt.strftime(date_format)  # Format the period before filtering
            key = (period, cook_name)
            aggregated_data.setdefault(key, []).append(time_taken)

        self.data_by_period = {}  # Store as a class attribute
        for (period, cook_name), times in aggregated_data.items():
            fastest_time = min(times)
            slowest_time = max(times)
            avg_time = int(sum(times) / len(times))
            ticket_count = len(times)
            self.data_by_period.setdefault(period, []).append(
                (cook_name, fastest_time, slowest_time, avg_time, ticket_count))

        # Sort records by avg_time for each period
        for period in self.data_by_period:
            self.data_by_period[period].sort(key=lambda x: x[3])  # Sort by avg_time (index 3 in tuple)

        # Display the sorted data
        for period, records in self.data_by_period.items():
            display_period = self.format_period(period, group_by)
            self.add_aggregated_grid_item(display_period, records)

    def format_period(self, period, group_by):
        """Format the period string for display."""
        if group_by == "hour":
            dt = datetime.strptime(period, "%Y-%m-%d %H:00")
            start_time = dt.strftime("%I%p").lstrip("0")  # Remove leading zero
            end_time = (dt + timedelta(hours=1)).strftime("%I%p").lstrip("0")  # Next hour
            return f"{dt.strftime('%b %d, %Y')} | {start_time} - {end_time}"
        elif group_by == "week":
            year, week = map(int, period.split('-'))
            first_day = datetime.strptime(f"{year}-W{week}-1", "%Y-W%W-%w")
            last_day = first_day + timedelta(days=6)
            return f"{first_day.strftime('%b %d, %Y')} - {last_day.strftime('%b %d, %Y')}"
        elif group_by == "month":
            year, month = map(int, period.split('-'))
            return f"{datetime(year, month, 1).strftime('%b %Y')}"
        elif group_by == "day":
            return datetime.strptime(period, "%Y-%m-%d").strftime('%a | %b %d, %Y')
        return period

    def add_aggregated_grid_item(self, period, records):
        """Add grid layout dynamically with proper spacing and sizing."""

        # Create a grid layout with sufficient spacing
        grid = GridLayout(cols=5, spacing=10, size_hint_y=None, padding=[20, 10, 20, 10])  # Add padding for grid
        grid.bind(minimum_height=grid.setter("height"))

        # Add headers to the grid
        headers = ["Cook:", "Shortest:", "Longest:", "Avg:", "Tickets:"]
        for header in headers:
            grid.add_widget(Label(text=header, bold=True, size_hint_y=None, height=40, underline=True, color=(0.506, 0.522, 0.565, 1)))  # Header height remains 40

        # Add records to the grid
        for cook_name, fastest, slowest, avg, count in records:
            grid.add_widget(Label(text=cook_name, size_hint_y=None, height=40, color=(0.894, 0.898, 0.914, 1)))
            grid.add_widget(Label(text=f"{fastest // 60}:{fastest % 60:02}", size_hint_y=None, height=40, color=(0.894, 0.898, 0.914, 1)))
            grid.add_widget(Label(text=f"{slowest // 60}:{slowest % 60:02}", size_hint_y=None, height=40, color=(0.894, 0.898, 0.914, 1)))
            grid.add_widget(Label(text=f"{avg // 60}:{avg % 60:02}", size_hint_y=None, height=40, color=(0.894, 0.898, 0.914, 1)))
            grid.add_widget(Label(text=str(count), size_hint_y=None, height=40, color=(0.894, 0.898, 0.914, 1)))

        # Force layout recalculation for the grid
        grid.do_layout()

        # Calculate the total height needed for the container
        total_height = grid.height + 100  # Add space for the period header (50) and padding

        # Create a container for the rounded box (transparent background)
        container = BoxLayout(orientation="vertical", size_hint_y=None)
        container.height = total_height

        # Create the rounded box with color and internal padding
        rounded_box_outter = RoundedBoxLayout(
            orientation="vertical",
            size_hint=(1, 1),  # Fills the container
            color=(0.247, 0.475, 0.424, 1),  # Use the desired color
            padding=2,
            spacing=2
        )

        # Create the rounded box with color and internal padding
        rounded_box = RoundedBoxLayout(
            orientation="vertical",
            size_hint=(1, 1),  # Fills the container
            color=(0.204, 0.408, 0.373, 1),  # Use the desired color
            padding=[30, 30, 30, 30],  # Add internal padding
        )

        # Add a properly sized period label
        rounded_box.add_widget(
            Label(text=period, size_hint_y=None, height=50, bold=True, valign="middle", underline=True, color=(0.714, 0.569, 0.129, 1)))  # Period Label
        rounded_box.add_widget(grid)

        # Add the rounded box to the container
        rounded_box_outter.add_widget(rounded_box)
        container.add_widget(rounded_box_outter)

        # Add the container to the main data container
        self.data_container.add_widget(container)
        self.data_container.do_layout()

    def go_back(self, instance):
        """Navigate back to the manager screen."""
        self.manager.current = "manager_screen"

    def export_to_excel(self, instance):
        """Export current performance data to an Excel file in the Documents folder with a pop-up notification."""

        def show_popup(message):
            """Displays a pop-up with the given message."""
            popup_layout = BoxLayout(orientation='vertical', spacing=10, padding=10)
            label = Label(text=message, size_hint_y=0.7, font_size=30, halign='center', valign='middle')
            dismiss_button = RoundedButton(text="Dismiss", size_hint_y=None, height=100,
                                           background_color=(0.541, 0.29, 0.29, 1), color=(0.894, 0.898, 0.914, 1),
                                           font_size=50)
            popup_layout.add_widget(label)
            popup_layout.add_widget(dismiss_button)

            popup = Popup(title="Export Status",
                          content=popup_layout,
                          title_size=50,
                          title_align='center',
                          size_hint=(0.8, 0.5),
                          separator_color=(0.169, 0.329, 0.298, 1),
                          )
            dismiss_button.bind(on_press=popup.dismiss)
            popup.open()

        if not hasattr(self, 'data_by_period') or not self.data_by_period:
            show_popup("No data available for export.")
            return

        def format_time(seconds):
            """Convert seconds to minute:seconds format."""
            minutes, seconds = divmod(seconds, 60)
            return f"{minutes}:{seconds:02d}"

        # ✅ Get correct local and UTC `start_date`
        local_tz = pytz.timezone("US/Eastern")  # Change to your actual timezone
        now = datetime.now(local_tz)

        if self.current_view == "daily":
            start_date = now - timedelta(days=7)
            date_format = "%Y-%m-%d"
        elif self.current_view == "weekly":
            start_date = now - timedelta(days=30)
            date_format = "%Y-%W"
        elif self.current_view == "monthly":
            start_date = now - timedelta(days=90)
            date_format = "%Y-%m"
        elif self.current_view == "hourly":
            start_date = now - timedelta(days=2)  # Last 2 days of hourly data
            date_format = "%Y-%m-%d %H:00"
        else:
            show_popup("Error: Unknown export range.")
            return

        start_date_utc = start_date.astimezone(pytz.utc).strftime("%Y-%m-%d %H:%M:%S")
        # Convert `start_date_utc` to a proper datetime object with UTC timezone
        start_date_utc = datetime.strptime(start_date_utc, "%Y-%m-%d %H:%M:%S").replace(tzinfo=pytz.utc)


        # ✅ Filter `data_by_period` using **correct** `start_date_utc`
        filtered_data = {}
        for period, records in self.data_by_period.items():
            try:
                # Convert stored period to datetime
                if self.current_view == "weekly":
                    year, week = map(int, period.split('-'))
                    period_date = datetime.strptime(f"{year}-W{week}-1", "%Y-W%W-%w")
                elif self.current_view == "monthly":
                    year, month = map(int, period.split('-'))
                    period_date = datetime(year, month, 1)
                else:
                    period_date = datetime.strptime(period, date_format)

                # Convert `period_date` to UTC to compare correctly
                period_date_utc = period_date.astimezone(pytz.utc)

                if period_date_utc >= start_date_utc:
                    filtered_data[period] = records
            except ValueError:
                continue

        if not filtered_data:
            show_popup("No data to export within the selected range.")
            return

        # ✅ Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Performance Data"

        # ✅ Add title
        start_date_str = start_date.strftime("%b %d, %Y")
        now_str = now.strftime("%b %d, %Y")
        title = f"Performance Data: ({start_date_str} - {now_str})"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = openpyxl.styles.Font(bold=True, size=16)
        title_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

        # ✅ Headers
        headers = ["Period", "Cook", "Shortest", "Longest", "Average", "Tickets"]
        ws.append(headers)
        for cell in ws[2]:
            cell.font = openpyxl.styles.Font(bold=True)

        # ✅ Add filtered data
        for period, records in filtered_data.items():
            for record in records:
                cook_name, fastest_time, slowest_time, avg_time, ticket_count = record
                ws.append([
                    period,
                    cook_name,
                    format_time(fastest_time),
                    format_time(slowest_time),
                    format_time(avg_time),
                    ticket_count
                ])

        # ✅ Determine Save Path
        try:
            documents_folder = storagepath.get_documents_dir()
            if not documents_folder:
                documents_folder = "/storage/emulated/0/Documents"  # ✅ Fallback for Android
        except Exception:
            documents_folder = "/storage/emulated/0/Documents"  # ✅ Another fallback

        # ✅ Generate file name based on current view
        start_date_file = start_date.strftime("%b%d_%Y")
        now_file = now.strftime("%b%d_%Y")
        file_name = f"Performance_{self.current_view.capitalize()}_{start_date_file}_to_{now_file}.xlsx"

        save_path = Path(documents_folder) / file_name  # Full path
        wb.save(save_path)

        # ✅ Success message
        show_popup(f"Saving to:\nFiles/Documents\n\nPlease allow a moment for it to be ready.")

    def on_leave(self, *args):
        """Reset the screen state when leaving."""
        # Clear the data container
        self.data_container.clear_widgets()

        # Reset button colors to default
        self.monthly_button.color_instruction.rgba = (0.204, 0.408, 0.373, 1)  # Default background color
        self.monthly_button.color = (0.714, 0.569, 0.129, 1)  # Default text color

        self.weekly_button.color_instruction.rgba = (0.204, 0.408, 0.373, 1)
        self.weekly_button.color = (0.714, 0.569, 0.129, 1)

        self.daily_button.color_instruction.rgba = (0.204, 0.408, 0.373, 1)
        self.daily_button.color = (0.714, 0.569, 0.129, 1)

        # Clear any stored data
        if hasattr(self, 'data_by_period'):
            del self.data_by_period
